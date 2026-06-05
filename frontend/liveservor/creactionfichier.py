# Import pour la créaction de mon fichier csv

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os
from pathlib import Path
    #Intégration de la fonction pour créer un fichier xlx
    
    
def enregistrer_presence(liste
):
    """
    Enregistre une présence dans une feuille correspondant à la date.

    Structure :

    Classeur Excel
    ├── 26-05-2026
    ├── 27-05-2026
    ├── 28-05-2026
    └── ...

    Chaque feuille contient un tableau Excel avec filtres.
    """
# ==========================================================
        # Ouvrir ou créer le classeur
        # ==========================================================


    wb = Workbook()

        # Suppression de la feuille par défaut
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    # ==========================================================
    # Date / Heure
    # ==========================================================
    for element in liste:
        source = element['source']
        nom_personne = element['nom']
        heure = element['heure']
        date_presence = element['date']

        nom_feuille = date_presence

        

        # ==========================================================
        # Création de la feuille du jour
        # ==========================================================

        if nom_feuille not in wb.sheetnames:

            ws = wb.create_sheet(title=nom_feuille)

            # Largeur colonnes
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 35
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 18

            # Entêtes
            headers = [
                "Source",
                "Noms et Prénoms",
                "Heures",
                "Date"
            ]

            for col, titre in enumerate(headers, start=1):
                cellule = ws.cell(row=1, column=col)
                cellule.value = titre

                cellule.font = Font(
                    name="Cambria",
                    size=12,
                    bold=True,
                    color="FFFFFF"
                )

                cellule.fill = PatternFill(
                    fill_type="solid",
                    fgColor="1800ad"
                )

                cellule.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # Figer l'entête
            ws.freeze_panes = "A2"

        else:
            ws = wb[nom_feuille]

        # ==========================================================
        # Vérification doublon
        # ==========================================================

        deja_present = False
        
        for row in ws.iter_rows(min_row=2, values_only=True):

            if row[0] == nom_personne:
                deja_present = True
                break

        if deja_present:
            print(f"{nom_personne} déjà enregistré aujourd'hui.")
            pass

        # ==========================================================
        # Ajout de la présence
        # ==========================================================

        ws.append([
            source,
            nom_personne,
            heure,
            date_presence
        ])

        # ==========================================================
        # Création / Mise à jour du tableau Excel
        # ==========================================================

        table_name = f"TableauPresence_{nom_feuille.replace('-', '_')}"

        if table_name in ws.tables:

            ws.tables[table_name].ref = f"A1:D{ws.max_row}"

        else:

            table = Table(
                displayName=table_name,
                ref=f"A1:D{ws.max_row}"
            )

            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            table.tableStyleInfo = style

            ws.add_table(table)

        # ==========================================================
        # Alignement
        # ==========================================================

        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=1,
            max_col=4
            ):
            
            for cell in row:
                cell.font =  Font(
                    name="Times New Roman",
                    size=11,
                    bold=True,
                    
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

    # ==========================================================
    # Sauvegarde Seulement qu'en django on ne sauvegarde plus 
    # ==========================================================
    print(f"Présence enregistrée")
    return wb

    

#Fin de cette fonction